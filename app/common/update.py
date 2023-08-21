from io import BytesIO
from openpyxl import load_workbook
import requests
from common.errors import message_box_fatal_error, warning_message
from common.constants import (
    GITHUB_CUSTOM_TRANSLATIONS_ZIP_URL,
    GITHUB_CLARITY_VERSION_UPDATE_URL,
    GITHUB_CLARITY_MERGE_XLSX_URL,
    GITHUB_CLARITY_MONSTERS_JSON_URL,
    GITHUB_CLARITY_NPC_JSON_URL,
    GITHUB_CLARITY_ITEMS_JSON_URL,
    GITHUB_CLARITY_KEY_ITEMS_JSON_URL,
    GITHUB_CLARITY_QUESTS_REQUESTS_JSON_URL,
    GITHUB_CLARITY_DAT1_URL,
    GITHUB_CLARITY_IDX_URL
)
from common.lib import get_abs_path, process_exists, check_if_running_as_admin
from loguru import logger
import os
import sqlite3
from subprocess import Popen
import sys
import winreg
from zipfile import ZipFile as zip
from common.translate import load_user_config, update_user_config
from tkinter.filedialog import askdirectory
import shutil


def download_custom_files():
    try:
        logger.info("Downloading custom json files.")
        request = requests.get(GITHUB_CUSTOM_TRANSLATIONS_ZIP_URL, timeout=15)
        if request.status_code == 200:
            zfile = zip(BytesIO(request.content))
            for obj in zfile.infolist():
                if obj.filename[-1] == "/":  # don't parse directories
                    continue
                obj.filename = os.path.basename(obj.filename)  # unzipped files copy zip folder structure, so re-assign filename to basename when we extract
                if obj.filename == "glossary.csv" or "custom_" in obj.filename:
                    zfile.extract(obj, "misc_files")
                if obj.filename in ["merge.xlsx"]:
                    zfile.extract(obj, "misc_files")

        for url in [
            GITHUB_CLARITY_MONSTERS_JSON_URL,
            GITHUB_CLARITY_NPC_JSON_URL,
            GITHUB_CLARITY_ITEMS_JSON_URL,
            GITHUB_CLARITY_KEY_ITEMS_JSON_URL,
            GITHUB_CLARITY_QUESTS_REQUESTS_JSON_URL
        ]:
            request = requests.get(url, timeout=15)
            if request.status_code == 200:
                misc_files = "/".join([get_abs_path(__file__), "../misc_files"])
                with open("/".join([misc_files, os.path.basename(url)]), "w+", encoding="utf-8") as f:
                    f.write(request.text)
        merge_local_db()
    except Exception as e:
        logger.error(f"Failed to download custom files. Error: {e}")
        input("Press ENTER to exit.")
        sys.exit()


def check_for_updates(update: bool):
    """
    Checks to see if Clarity is running the latest version of itself.
    If not, will launch updater.py and exit.

    :param update: Whether or not to update after checking for updates.
    """
    logger.info("Checking dqxclarity repo for updates...")
    if not os.path.exists("version.update"):
        logger.warning("Couldn't determine current version of dqxclarity. Running as is.")
        return

    with open("version.update", "r") as file:
        cur_ver = file.read().strip()

    try:
        url = GITHUB_CLARITY_VERSION_UPDATE_URL
        github_request = requests.get(url)
    except requests.exceptions.RequestException as e:
        logger.warning(f"Failed to check latest version. Running anyways.\n{e}")
        return

    try:
        release_version = github_request.json()["tag_name"]
        if release_version.startswith("v"):
            release_version = release_version[1:]
        if release_version == cur_ver:
            logger.success(f"Clarity is up to date! (Current version: {str(cur_ver)})")
        else:
            logger.warning(f"Clarity is out of date! (Current: {str(cur_ver)}, Latest: {str(release_version)}).")
            if update:
                install_path = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Python\PythonCore\3.11-32\InstallPath")
                python_exe = winreg.QueryValueEx(install_path, "ExecutablePath")
                if not python_exe:
                    logger.warning("Did not find Python exe! Clarity is unable to update and will continue without updating.")
                    return False
                logger.info(f"Launching updater.")
                Popen([python_exe[0], "./updater.py"])
                sys.exit()
        return
    except Exception as e:
        logger.warning(f"There was a problem checking trying to update. Clarity will continue without updating.\n{e}")
        return


def merge_local_db():
    merge_file = "/".join([get_abs_path(__file__), "../misc_files/merge.xlsx"])
    db_file = "/".join([get_abs_path(__file__), "../misc_files/clarity_dialog.db"])

    records_inserted = 0
    records_updated = 0

    if os.path.exists(merge_file):
        os.remove(merge_file)

    try:
        url = GITHUB_CLARITY_MERGE_XLSX_URL
        r = requests.get(url, timeout=15)
    except Exception as e:
        message_box_fatal_error("Timeout", str(e))

    with open(merge_file, "wb") as merge:
        merge.write(r.content)
        logger.info("Local database file downloaded.")

    if os.path.exists(merge_file):
        wb = load_workbook(merge_file)
        ws_dialogue = wb["Dialogue"]
        ws_walkthrough = wb["Walkthrough"]
        ws_quests = wb["Quests"]

        ###Dialogue insertion
        for rowNum in range(2, ws_dialogue.max_row + 1):
            source_text = ws_dialogue.cell(row=rowNum, column=1).value
            en_text = ws_dialogue.cell(row=rowNum, column=3).value

            escaped_text = en_text.replace("'", "''")
            table = "dialog"
            npc_name = ""
            language = "en"
            # bad_string = "魔物の軍団は　撤退していった。"
            bad_string = False
            bad_string_col = str(ws_dialogue.cell(row=rowNum, column=4).value)
            if "BAD STRING" in bad_string_col:
                bad_string = True

            try:
                conn = sqlite3.connect(db_file)
                if bad_string:
                    selectQuery = f"SELECT ja FROM dialog WHERE ja LIKE '%{source_text}%'"
                else:
                    selectQuery = f"SELECT ja FROM dialog WHERE ja = '{source_text}'"
                if bad_string:
                    updateQuery = f"UPDATE dialog SET en = '{escaped_text}' WHERE ja LIKE '%{source_text}%'"
                else:
                    updateQuery = f"UPDATE dialog SET en = '{escaped_text}' WHERE ja = '{source_text}'"
                if not bad_string:
                    insertQuery = f"INSERT INTO dialog (ja, npc_name, en) VALUES ('{source_text}', '{npc_name}', '{escaped_text}')"
                else:
                    insertQuery = ""

                cursor = conn.cursor()
                results = cursor.execute(selectQuery)

                if results.fetchone() is None and insertQuery != "":
                    cursor.execute(insertQuery)
                    records_inserted += 1
                else:
                    cursor.execute(updateQuery)
                    records_updated += 1

                conn.commit()
                cursor.close()
            except sqlite3.Error as e:
                raise Exception(f"Unable to write data to table: {e}")
            finally:
                if conn:
                    conn.close()

        # Walkthrough insertion
        for rowNum in range(2, ws_walkthrough.max_row + 1):
            source_text = ws_walkthrough.cell(row=rowNum, column=1).value
            en_text = ws_walkthrough.cell(row=rowNum, column=3).value

            escaped_text = en_text.replace("'", "''")

            try:
                conn = sqlite3.connect(db_file)
                selectQuery = f"SELECT ja FROM walkthrough WHERE ja = '{source_text}'"

                updateQuery = f"UPDATE walkthrough SET en = '{escaped_text}' WHERE ja = '{source_text}'"
                insertQuery = f"INSERT INTO walkthrough (ja, en) VALUES ('{source_text}', '{escaped_text}')"

                cursor = conn.cursor()
                results = cursor.execute(selectQuery)

                if results.fetchone() is None:
                    cursor.execute(insertQuery)
                    records_inserted += 1
                else:
                    cursor.execute(updateQuery)
                    records_updated += 1

                conn.commit()
                cursor.close()
            except sqlite3.Error as e:
                raise Exception(f"Unable to write data to table: {e}")
            finally:
                if conn:
                    conn.close()
                    
        # Quests insertion
        for rowNum in range(2, ws_quests.max_row + 1):
            source_text = ws_quests.cell(row=rowNum, column=1).value
            en_text = ws_quests.cell(row=rowNum, column=3).value

            escaped_text = en_text.replace("'", "''")
            
            bad_string = False
            bad_string_col = str(ws_quests.cell(row=rowNum, column=4).value)
            if "BAD STRING" in bad_string_col:
                bad_string = True

            try:
                conn = sqlite3.connect(db_file)
                if bad_string:
                    selectQuery = f"SELECT ja FROM quests WHERE ja LIKE '%{source_text}%'"
                    updateQuery = f"UPDATE quests SET en = '{escaped_text}' WHERE ja LIKE '%{source_text}%'"
                else:
                    selectQuery = f"SELECT ja FROM quests WHERE ja = '{source_text}'"
                    updateQuery = f"UPDATE quests SET en = '{escaped_text}' WHERE ja = '{source_text}'"
                    
                if not bad_string:
                    insertQuery = f"INSERT INTO quests (ja, en) VALUES ('{source_text}', '{escaped_text}')"
                else:
                    insertQuery = ""

                cursor = conn.cursor()
                results = cursor.execute(selectQuery)

                if results.fetchone() is None:
                    cursor.execute(insertQuery)
                    records_inserted += 1
                else:
                    cursor.execute(updateQuery)
                    records_updated += 1

                conn.commit()
                cursor.close()
            except sqlite3.Error as e:
                raise Exception(f"Unable to write data to table: {e}")
            finally:
                if conn:
                    conn.close()

        logger.info(str(records_inserted) + " records were inserted into local db.")
        logger.info(str(records_updated) + " records in local db were updated.")
        

def download_dat_files():
    """
    Verifies the user's DQX install location and prompts
    them to locate it if not found. Uses this location to
    download the latest data files from the dqxclarity repo.
    """
    if process_exists("DQXGame.exe"):
        message = "Please close DQX before attempting to update the translated DAT/IDX file."
        logger.error(message)
        message_box_fatal_error(
            title="DQXGame.exe is open",
            message=message
        )

    if not check_if_running_as_admin():
        message = "dqxclarity must be running as an administrator in order to update the translated DAT/IDX file. Please re-launch dqxclarity as an administrator and try again."
        logger.error(message)
        message_box_fatal_error(
            title="Program not elevated",
            message=message
        )

    config = load_user_config()
    dat0_file = "data00000000.win32.dat0"
    idx0_file = "data00000000.win32.idx"

    install_directory = config["config"]["installdirectory"]

    valid_path = False
    if install_directory:
        if os.path.isdir(install_directory):
            logger.success("DQX game path is valid.")
            valid_path = True

    if not valid_path:
        default_path = 'C:/Program Files (x86)/SquareEnix/DRAGON QUEST X'
        if os.path.exists(default_path):
            update_user_config('config', 'installdirectory', default_path)
        else:
            warning_message(
                title="[dqxclarity] Couldn't Find DQX Directory",
                message="Could not find DQX directory. Browse to the path where you installed the game and select the \"DRAGON QUEST X\" folder."
            )

            while True:
                dqx_path = askdirectory()
                dat0_path = "/".join([dqx_path, "Game/Content/Data", dat0_file])

                if os.path.isfile(dat0_path):
                    update_user_config('config', 'installdirectory', dqx_path)
                    logger.success("DQX path verified.")
                    break
                else:
                    warning_message(
                        title="[dqxclarity] Invalid Directory",
                        message="The path you provided is not a valid DQX path.\nBrowse to the path where you installed the game and select the \"DRAGON QUEST X\" folder."
                    )

    config = load_user_config()  # call this again in case we made changes above
    dqx_path = "/".join([config['config']['installdirectory'], "Game/Content/Data"]) 
    idx0_path = "/".join([dqx_path, idx0_file])

    if not os.path.isfile(f"{idx0_path}.bak"):
        logger.info(f"Did not find a backup of existing idx file. Backing up and renaming to {idx0_file}.bak")
        shutil.copy(idx0_path, f"{idx0_path}.bak")

    try:
        logger.info("Downloading DAT1 and IDX files.")
        dat_request = requests.get(GITHUB_CLARITY_DAT1_URL, timeout=10)
        idx_request = requests.get(GITHUB_CLARITY_IDX_URL, timeout=10)

        # Make sure both requests are good before we write the files
        if dat_request.status_code == 200 and idx_request.status_code == 200:
            with open(dqx_path + "/data00000000.win32.dat1", "w+b") as f:
                f.write(dat_request.content)
            with open(dqx_path + "/data00000000.win32.idx", "w+b") as f:
                f.write(idx_request.content)
            logger.success("Translation files downloaded.")
        else:
            logger.error("Failed to download translation files. Clarity will continue without updating translation files")
    except Exception as e:
        logger.error(f"Failed to download data files. Error: {e}")
